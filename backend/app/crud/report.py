"""생산 보고서 CRUD 모듈."""

from datetime import date, timedelta
from typing import Any

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.equipment_ext import EquipmentFailure
from app.models.production import QCRecord, WorkOrder, WorkOrderResult


class CRUDReport:
    """생산 보고서 CRUD 클래스."""

    def _aggregate_production(
        self,
        db: Session,
        *,
        date_from: date,
        date_to: date,
    ) -> dict[str, Any]:
        """지정 기간의 생산 실적을 집계합니다.

        Args:
            db: 데이터베이스 세션
            date_from: 시작일
            date_to: 종료일

        Returns:
            집계 결과 딕셔너리
        """
        from app.models.product import Product

        # 제품별 집계
        rows = (
            db.query(
                Product.product_name,
                func.sum(WorkOrder.planned_qty).label("planned_qty"),
                func.sum(WorkOrder.actual_qty).label("actual_qty"),
                func.sum(WorkOrder.defect_qty).label("defect_qty"),
            )
            .join(Product, Product.id == WorkOrder.product_id)
            .filter(
                WorkOrder.is_deleted == False,
                WorkOrder.status == "COMPLETED",
                func.date(WorkOrder.work_date) >= date_from,
                func.date(WorkOrder.work_date) <= date_to,
            )
            .group_by(Product.product_name)
            .all()
        )

        by_product = []
        total_planned = 0
        total_actual = 0
        total_defect = 0

        for row in rows:
            planned = int(row.planned_qty or 0)
            actual = int(row.actual_qty or 0)
            defect = int(row.defect_qty or 0)
            total_planned += planned
            total_actual += actual
            total_defect += defect

            achievement_rate = round(actual / planned * 100, 2) if planned > 0 else 0.0
            defect_rate = round(defect / actual * 100, 2) if actual > 0 else 0.0

            by_product.append(
                {
                    "product_name": row.product_name,
                    "planned_qty": planned,
                    "actual_qty": actual,
                    "defect_qty": defect,
                    "achievement_rate": achievement_rate,
                    "defect_rate": defect_rate,
                }
            )

        overall_achievement = round(total_actual / total_planned * 100, 2) if total_planned > 0 else 0.0
        overall_defect_rate = round(total_defect / total_actual * 100, 2) if total_actual > 0 else 0.0

        # CCP 이탈 건수
        ccp_violations = (
            db.query(func.count(QCRecord.id))
            .filter(
                QCRecord.is_deleted == False,
                QCRecord.is_pass == False,
                func.date(QCRecord.created_at) >= date_from,
                func.date(QCRecord.created_at) <= date_to,
            )
            .scalar()
        ) or 0

        # 설비 비가동 시간 (분 단위 환산)
        failures = (
            db.query(EquipmentFailure)
            .filter(
                EquipmentFailure.is_deleted == False,
                func.date(EquipmentFailure.failure_date) >= date_from,
                func.date(EquipmentFailure.failure_date) <= date_to,
                EquipmentFailure.downtime_hours.isnot(None),
            )
            .all()
        )
        equipment_downtime_minutes = int(
            sum(float(f.downtime_hours or 0) * 60 for f in failures)
        )

        return {
            "total_planned": total_planned,
            "total_actual": total_actual,
            "total_defect": total_defect,
            "overall_achievement": overall_achievement,
            "overall_defect_rate": overall_defect_rate,
            "by_product": by_product,
            "ccp_violations": ccp_violations,
            "equipment_downtime_minutes": equipment_downtime_minutes,
        }

    def get_daily(self, db: Session, *, target_date: date) -> dict[str, Any]:
        """일별 보고서를 반환합니다.

        Args:
            db: 데이터베이스 세션
            target_date: 조회 일자

        Returns:
            보고서 딕셔너리
        """
        agg = self._aggregate_production(db, date_from=target_date, date_to=target_date)
        return {
            "period": "daily",
            "date_from": str(target_date),
            "date_to": str(target_date),
            **agg,
        }

    def get_weekly(self, db: Session, *, week_start: date) -> dict[str, Any]:
        """주별 보고서를 반환합니다.

        Args:
            db: 데이터베이스 세션
            week_start: 주 시작일 (월요일)

        Returns:
            보고서 딕셔너리
        """
        week_end = week_start + timedelta(days=6)
        agg = self._aggregate_production(db, date_from=week_start, date_to=week_end)
        return {
            "period": "weekly",
            "date_from": str(week_start),
            "date_to": str(week_end),
            **agg,
        }

    def get_monthly(self, db: Session, *, year: int, month: int) -> dict[str, Any]:
        """월별 보고서를 반환합니다.

        Args:
            db: 데이터베이스 세션
            year: 연도
            month: 월

        Returns:
            보고서 딕셔너리
        """
        from calendar import monthrange

        _, last_day = monthrange(year, month)
        date_from = date(year, month, 1)
        date_to = date(year, month, last_day)

        agg = self._aggregate_production(db, date_from=date_from, date_to=date_to)
        return {
            "period": "monthly",
            "date_from": str(date_from),
            "date_to": str(date_to),
            **agg,
        }

    def export_excel(self, db: Session, *, report_type: str, **kwargs) -> bytes:
        """보고서를 Excel 파일로 내보냅니다.

        openpyxl을 사용하여 헤더 + 데이터 + 총계 행을 포함한
        Excel 파일을 bytes로 반환합니다.

        Args:
            db: 데이터베이스 세션
            report_type: 보고서 유형 (daily / weekly / monthly)
            **kwargs: 보고서 파라미터 (date, week_start, year, month)

        Returns:
            Excel 파일 bytes
        """
        import io

        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        # 보고서 데이터 가져오기
        if report_type == "daily":
            target_date = kwargs.get("date", date.today())
            report = self.get_daily(db, target_date=target_date)
        elif report_type == "weekly":
            week_start = kwargs.get("week_start", date.today())
            report = self.get_weekly(db, week_start=week_start)
        elif report_type == "monthly":
            year = kwargs.get("year", date.today().year)
            month = kwargs.get("month", date.today().month)
            report = self.get_monthly(db, year=year, month=month)
        else:
            report = self.get_daily(db, target_date=date.today())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "생산보고서"

        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # 보고서 정보 헤더
        ws["A1"] = "임진강김치(주) 생산보고서"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"기간: {report['date_from']} ~ {report['date_to']}"
        ws["A3"] = f"종류: {report['period'].upper()}"
        ws.append([])

        # 컬럼 헤더
        headers = ["제품명", "계획수량", "실적수량", "불량수량", "달성률(%)", "불량률(%)"]
        ws.append(headers)
        header_row = ws.max_row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 제품별 데이터
        for row_data in report.get("by_product", []):
            ws.append([
                row_data["product_name"],
                row_data["planned_qty"],
                row_data["actual_qty"],
                row_data["defect_qty"],
                row_data["achievement_rate"],
                row_data["defect_rate"],
            ])

        # 총계 행
        total_row = [
            "합계",
            report["total_planned"],
            report["total_actual"],
            report["total_defect"],
            report["overall_achievement"],
            report["overall_defect_rate"],
        ]
        ws.append(total_row)
        total_row_idx = ws.max_row
        for col_idx in range(1, len(total_row) + 1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = total_fill

        ws.append([])

        # 요약 정보
        ws.append(["CCP 이탈 건수", report["ccp_violations"]])
        ws.append(["설비 비가동 시간(분)", report["equipment_downtime_minutes"]])

        # 열 너비 자동 조정
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            adjusted_width = max_length + 4
            ws.column_dimensions[col_letter].width = adjusted_width

        # bytes 변환
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()


crud_report = CRUDReport()
